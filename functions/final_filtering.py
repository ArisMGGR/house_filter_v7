from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#--------------FUNCTIONS--------------------------------------------------#


def M_create_excell(spreadsheet,sheet_extracted_page_name,categories_sheets):
    
    rest_of_numbers = get_all_extracted_numbers_sheets(spreadsheet,sheet_extracted_page_name)

    numbers_found,rest_of_numbers = get_all_sheet_matching_numbers(categories_sheets,spreadsheet,rest_of_numbers)

    categories_sheets = [x for x in categories_sheets if x not in ("NSA", "OTHER")]     #REMOVES NSA & OTHER CATEGORY

    user_interfase_for_category(categories_sheets)
    
    user_input_filtering = input("\t")
    
    numbers_found = categories_filtering_n_results(user_input_filtering,categories_sheets,numbers_found)
    
    numbers_found = interactive_delete(numbers_found)
        
    for item in numbers_found:
        rest_of_numbers.append([item[0],item[1]])

    create_whatsapp_excel(rest_of_numbers)


def get_all_extracted_numbers_sheets(spreadsheet,sheet_extracted_page_name):
    worksheet = spreadsheet.worksheet(sheet_extracted_page_name)
    return worksheet.get_all_values()


def filter_with_sheets(sheet_name,spreadsheet,curr_number_list):
    worksheet = spreadsheet.worksheet(sheet_name)
    sheet_numbers_list = worksheet.get_all_values()
    num_match = []
    no_num_match = []
    #print(sheet_numbers_list)
    if(sheet_numbers_list!=[[]]):
        for number_from_list in curr_number_list:
            num_found = False
            for number_from_sheets in sheet_numbers_list:
                if(number_from_sheets[0]==number_from_list[0] and num_found==False):
                    num_match.append([number_from_list[0],number_from_list[1],number_from_sheets[1]])
                    num_found = True
            if num_found==False:
                no_num_match.append(number_from_list)
    else:
        no_num_match = curr_number_list
    return num_match,no_num_match


def final_filter_with_sheet_data(numbers_found,category):
    temp_number_f = []
    counter = 0
    for number_item in numbers_found:
        if not (category in number_item[2]):
            temp_number_f.append(number_item)
        else:
            counter+=1
    print(f"\t+\t     {category} -> {counter}\t\t+")
    return temp_number_f


def final_filter_with_sheet_data2(numbers_found, category):
    temp_number_f = []
    counter = 0
    for number_item in numbers_found:
        if category in number_item[2].split():  # checks word by word
            temp_number_f.append(number_item)
        elif category in number_item[2]:  # substring match
            temp_number_f.append(number_item)
        else:
            counter += 1
    print(f"\t+\t     {category} -> {counter}\t\t+")
    return temp_number_f


def final_filter_with_sheet_data_general(numbers_found,category):
    temp_number_f = []
    for number_item in numbers_found:
        if(number_item[2]==category):
            temp_number_f.append(number_item)
    return temp_number_f


def interactive_delete(records):
    while records:
        print("\n+---------------------------------------+")
        for item in records:
            print(f"+ {item}")
        print("+---------------------------------------+")
        user_input = input("Enter phone number to delete (or 'y' to stop): ").strip()
        
        if user_input.lower() == "y":
            print("\nStopping...")
            break

        # try to delete by phone number
        found = False
        for r in records:
            if r[0] == user_input:
                records.remove(r)
                print(f"\nDeleted record {user_input}")
                found = True
                break
        
        if not found:
            print(f"\nNo record found {user_input}\n")
        
        
    
    if not records:
        print("\nStopping...")
    
    return records


def filter_categories(user_input, categories):
    # convert input digits into a set of integers
    try:
        to_remove = {int(ch) for ch in user_input}
    except ValueError:
        return categories[:]  # if input has non-digits, just return unchanged
    
    # build a new list excluding those indices
    return [cat for i, cat in enumerate(categories) if i not in to_remove]


def create_whatsapp_excel(data):
    filename = 'output.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "WhatsApp Contacts"

    # === HEADER ROW (Row 1) ===
    ws["A1"] = "WhatsApp Number(with country code)"
    ws["B1"] = "First Name"
    ws["C1"] = "Last Name"
    ws["D1"] = "Other"
    ws["F1"] = "Tips"

    # === SECOND ROW (Row 2) ===
    ws["B2"] = "Sender"
    ws["C2"] = "WAPI"
    ws["D2"] = "WAPI Sender Support"

    # === TIPS SECTION (starting from F3) ===
    tips = [
        "1.WhatsApp Number is required, please fill in the format：",
        "Country code in front, like： +19197646821,+8613119140503",
        "",
        "2.Please note that:",
        "[Required] The first column has to be the contact numbers.",
        "[Not required]Other columns can add customisation from excel."
    ]
    for i, tip in enumerate(tips, start=3):
        ws[f"F{i}"] = tip

    # === Insert Data Starting from Row 3 ===
    row_start = 3
    for i, entry in enumerate(data):
        number, name_str = entry
        full_name = name_str.replace("/", " ")

        row = row_start + i
        ws[f"A{row}"] = number
        ws[f"B{row}"] = full_name
        # Columns C and D are intentionally left blank

    # === Auto-adjust column widths ===
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2  # Add padding
        ws.column_dimensions[col_letter].width = adjusted_width

    # === Save workbook ===
    
    wb.save(filename)
    print("\n")
    print("\t+------------------------------------+")
    print(f"\t+\t  Total Entries - {len(data)}\t     +")
    print(f"\t+  Excel file '{filename}' created  +")
    print("\t+------------------------------------+\n")


def categories_filtering_n_results(user_input_filtering,G_categories_sheets,numbers_found):
    print("\t+-------------------------------+")
        
    numbers_found = final_filter_with_sheet_data(numbers_found,"NSA")
    G_categories_sheets = filter_categories(user_input_filtering,G_categories_sheets)
    for category in G_categories_sheets:
        numbers_found = final_filter_with_sheet_data(numbers_found,category)

    print("\t+-------------------------------+\n\n")
    return numbers_found


def user_interfase_for_category(G_categories_sheets):
    print("\n\t+-------------------------------+")
    i=0
    for category in G_categories_sheets:
        if(len(category)==3):
            print(f"\t+\t  {i} -> add {category}\t\t+")
        elif(len(category)==4):
            print(f"\t+\t  {i} -> add {category}\t\t+")
        else:
            print(f"\t+\t  {i} -> add {category}               +")
        i+=1
    print("\t+-------------------------------+\n")


def get_all_sheet_matching_numbers(categories_sheets,spreadsheet,rest_of_numbers):
    numbers_found = []           #this will hold the numbers that are both in filtered_numbers and the google sheet
    for current_sheet in categories_sheets:
        temp_numbers_found,t_rest_of_numbers = filter_with_sheets(current_sheet,spreadsheet,rest_of_numbers)
        rest_of_numbers = t_rest_of_numbers
        for item in temp_numbers_found:
            numbers_found.append(item)

    return numbers_found,rest_of_numbers